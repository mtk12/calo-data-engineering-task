import os

import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.drawing.image import Image


def write_to_sheet(df, sheet_name, file_path):
    for col in df.select_dtypes(include=['datetimetz']).columns:
        df[col] = df[col].dt.tz_localize(None)

    if os.path.exists(file_path):
        with pd.ExcelWriter(file_path, engine='openpyxl', mode='a', if_sheet_exists='replace') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    else:
        with pd.ExcelWriter(file_path, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)


def generate_excel(parsed_df, error_df, analysis_df, output_excel_file_name, output_excel_file_path):
    full_path = os.path.join(output_excel_file_path, output_excel_file_name)

    write_to_sheet(analysis_df, "user_wise_error_analysis_data", full_path)
    write_to_sheet(parsed_df, "transaction_data", full_path)
    write_to_sheet(error_df, "error_data", full_path)


def insert_chart_to_excel(output_excel_file_path, output_excel_file_name, sheet_name, image_path, cell):
    full_path = os.path.join(output_excel_file_path, output_excel_file_name)
    if os.path.exists(full_path):
        workbook = load_workbook(full_path)
    else:
        workbook = Workbook()

    if sheet_name in workbook.sheetnames:
        worksheet = workbook[sheet_name]
    else:
        worksheet = workbook.create_sheet(sheet_name)

    img = Image(image_path)
    worksheet.add_image(img, cell)
    workbook.save(full_path)
